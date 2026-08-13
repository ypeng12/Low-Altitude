#!/usr/bin/env python3
"""Format XLSX files in data/analyze/ with professional openpyxl styling (auto-column widths, header styling, zebra striping, freeze panes)."""

import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

analyze_dir = Path("data/analyze")

xlsx_files = [
    analyze_dir / "gold_emotion_master.xlsx",
    analyze_dir / "gold_emotion_nrc_combined.xlsx"
]

header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

for file_path in xlsx_files:
    if not file_path.exists():
        continue
        
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Header styling
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows styling
    for row_idx in range(2, ws.max_row + 1):
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)
            
            col_name = str(ws.cell(row=1, column=col_idx).value).lower()
            if 'freq' in col_name or 'count' in col_name:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in ('word', 'canonical_lemma', 'in_nrc_lexicon'):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        col_name = str(col[0].value)
        
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1:
                max_len = max(max_len, len(val_str))
            else:
                # Limit example_context column width to max 50
                if 'example_context' in col_name:
                    max_len = 50
                    break
                max_len = max(max_len, min(len(val_str), 45))
                
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(file_path)
    print(f"Successfully formatted professional XLSX: {file_path.name}")

print("\nAll XLSX files in data/analyze/ formatted cleanly!")
